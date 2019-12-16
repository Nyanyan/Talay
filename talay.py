import serial
import serial.tools.list_ports
import os
import tkinter
import openpyxl

def search_com_port():
    coms = serial.tools.list_ports.comports()
    comlist = []
    for com in coms:
        comlist.append(com.device)
    print('Connected COM ports: ' + str(comlist))
    use_port = comlist[0]
    print('Use COM port: ' + use_port)
    return use_port

def findmaxrow(n):
    global maxrow
    for i in range(maxrow[n], sheet.max_row + 1):
        if sheet.cell(row = i, column = 1).value == None and maxrow[1-n] != i:
            return i
    return 0

def isNum(string):
    arr = ['0','1','2','3','4','5','6','7','8','9','.']
    ans = True
    for i in range(len(string)):
        if not(string[i] in arr):
            ans = False
    return ans

def nameinput(n):
    global maxrow
    row = 0
    if n == 0:
        name = name0box.get()
    else:
        name = name1box.get()
    sheet.cell(row=maxrow[n], column=1, value=name)
    if n == 0:
        name0box.config(state = 'disable')
        name0entbutton.config(state = 'disable')
    else:
        name1box.config(state = 'disable')
        name1entbutton.config(state = 'disable')

def timeinput(turn, n):
    global maxrow
    row = 0
    if n == 0:
        row = maxrow[0]
    else:
        row = maxrow[1]
    if n == 0:
        if turn == 0:
            t10box.config(state='disable')
            t10entbutton.config(state='disable')
            time[n][turn] = t10box.get()
        if turn == 1:
            t20box.config(state='disable')
            t20entbutton.config(state='disable')
            time[n][turn] = t20box.get()
        if turn == 2:
            t30box.config(state='disable')
            t30entbutton.config(state='disable')
            time[n][turn] = t30box.get()
        if turn == 3:
            t40box.config(state='disable')
            t40entbutton.config(state='disable')
            time[n][turn] = t40box.get()
        if turn == 4:
            t50box.config(state='disable')
            t50entbutton.config(state='disable')
            time[n][turn] = t50box.get()
    elif n == 1:
        if turn == 0:
            t11box.config(state='disable')
            t11entbutton.config(state='disable')
            time[n][turn] = t11box.get()
        if turn == 1:
            t21box.config(state='disable')
            t21entbutton.config(state='disable')
            time[n][turn] = t21box.get()
        if turn == 2:
            t31box.config(state='disable')
            t31entbutton.config(state='disable')
            time[n][turn] = t31box.get()
        if turn == 3:
            t41box.config(state='disable')
            t41entbutton.config(state='disable')
            time[n][turn] = t41box.get()
        if turn == 4:
            t51box.config(state='disable')
            t51entbutton.config(state='disable')
            time[n][turn] = t51box.get()
    sheet.cell(row=row, column=turn + 2, value=time[n][turn])
    wb.save(workbook)


def confirmtime():
    t10box.delete(0, tkinter.END)
    t10box.insert(tkinter.END,time[0][0])
    t20box.delete(0, tkinter.END)
    t20box.insert(tkinter.END,time[0][1])
    t30box.delete(0, tkinter.END)
    t30box.insert(tkinter.END,time[0][2])
    t40box.delete(0, tkinter.END)
    t40box.insert(tkinter.END,time[0][3])
    t50box.delete(0, tkinter.END)
    t50box.insert(tkinter.END,time[0][4])
    t11box.delete(0, tkinter.END)
    t11box.insert(tkinter.END,time[1][0])
    t21box.delete(0, tkinter.END)
    t21box.insert(tkinter.END,time[1][1])
    t31box.delete(0, tkinter.END)
    t31box.insert(tkinter.END,time[1][2])
    t41box.delete(0, tkinter.END)
    t41box.insert(tkinter.END,time[1][3])
    t51box.delete(0, tkinter.END)
    t51box.insert(tkinter.END,time[1][4])

def nextperson(n):
    global maxrow, time, turn

    if n == 0:
        time[0][0] = t10box.get()
        time[0][1] = t20box.get()
        time[0][2] = t30box.get()
        time[0][3] = t40box.get()
        time[0][4] = t50box.get()
        name = name0box.get()
    else:
        time[1][0] = t11box.get()
        time[1][1] = t21box.get()
        time[1][2] = t31box.get()
        time[1][3] = t41box.get()
        time[1][4] = t51box.get()
        name = name1box.get()

    flag = False
    if name == '':
        flag = True
    else:
        for i in range(5):
            if time[n][i] == '':
                flag = True
                break
    if flag:
        warning = tkinter.Tk()
        warning.title("Warning")
        warning.geometry("200x100")
        warninglabel = tkinter.Label(warning,text='Input Name or Time')
        warninglabel.pack()
        breakbutton = tkinter.Button(warning, text='OK', command=warning.destroy)
        breakbutton.pack()
        return 0
    
    tim = []
    DNFcnt = 0
    for i in range(5):
        if isNum(time[n][i]):
            tim.append(float(time[n][i]))
        else:
            DNFcnt += 1
    if DNFcnt == 0:
        average = round((sum(tim) - max(tim) - min(tim)) / 3, 3)
    elif DNFcnt == 1:
        average = round((sum(tim) - min(tim)) / 3, 3)
    else:
        average = 'DNF'

    sheet.cell(row=maxrow[n], column=1, value=name)
    for i in range(5):
        if isNum(time[n][i]):
            sheet.cell(row=maxrow[n] + 1, column=i + 2, value=float(time[n][i]))
        else:
            sheet.cell(row=maxrow[n] + 1, column=i + 2, value=time[n][i])
    sheet.cell(row=maxrow[n], column=7, value=average)

    wb.save(workbook)

    window = tkinter.Tk()
    window.title("Average")
    window.geometry("200x100")
    windowlabel = tkinter.Label(window,text='Your average is ' + str(average))
    windowlabel.pack()
    breakbutton = tkinter.Button(window, text='OK', command=window.destroy)
    breakbutton.pack()

    if n == 0:
        name0box.config(state='normal')
        name0entbutton.config(state='normal')

        t10box.config(state='normal')
        t10entbutton.config(state='normal')
        t20box.config(state='normal')
        t20entbutton.config(state='normal')
        t30box.config(state='normal')
        t30entbutton.config(state='normal')
        t40box.config(state='normal')
        t40entbutton.config(state='normal')
        t50box.config(state='normal')
        t50entbutton.config(state='normal')

        name0box.delete(0, tkinter.END)
        t10box.delete(0, tkinter.END)
        t20box.delete(0, tkinter.END)
        t30box.delete(0, tkinter.END)
        t40box.delete(0, tkinter.END)
        t50box.delete(0, tkinter.END)
    else:
        name1box.config(state='normal')
        name1entbutton.config(state='normal')

        t11box.config(state='normal')
        t11entbutton.config(state='normal')
        t21box.config(state='normal')
        t21entbutton.config(state='normal')
        t31box.config(state='normal')
        t31entbutton.config(state='normal')
        t41box.config(state='normal')
        t41entbutton.config(state='normal')
        t51box.config(state='normal')
        t51entbutton.config(state='normal')

        name1box.delete(0, tkinter.END)
        t11box.delete(0, tkinter.END)
        t21box.delete(0, tkinter.END)
        t31box.delete(0, tkinter.END)
        t41box.delete(0, tkinter.END)
        t51box.delete(0, tkinter.END)

    for i in range(5):
        time[n][i] = ''
    turn[n] = 0
    maxrow[n] = findmaxrow(n)

def turncount():
    global turn
    if t10box.get() == '':
        turn[0] = 0
    elif t20box.get() == '':
        turn[0] = 1
    elif t30box.get() == '':
        turn[0] = 2
    elif t40box.get() == '':
        turn[0] = 3
    elif t50box.get() == '':
        turn[0] = 4

    if t11box.get() == '':
        turn[1] = 0
    elif t21box.get() == '':
        turn[1] = 1
    elif t31box.get() == '':
        turn[1] = 2
    elif t41box.get() == '':
        turn[1] = 3
    elif t51box.get() == '':
        turn[1] = 4

def inputserial():
    global time, turn
    time[0][0] = t10box.get()
    time[0][1] = t20box.get()
    time[0][2] = t30box.get()
    time[0][3] = t40box.get()
    time[0][4] = t50box.get()
    time[1][0] = t11box.get()
    time[1][1] = t21box.get()
    time[1][2] = t31box.get()
    time[1][3] = t41box.get()
    time[1][4] = t51box.get()
    turncount()
    ser = serial.Serial()
    ser.port = port
    ser.baundrate = 9600
    ser.timeout = 0.1 #sec
    ser.setDTR(False)
    ser.open()
    #print(turn)
    line = ser.readline().decode('utf8', 'ignore').rstrip(os.linesep)
    if line != '':
        #print(line)
        f = False
        arr = ['0','1']
        if line[0] in arr:
            if len(line) == 10:
                checksum = 64
                for j in range(6):
                    tmp = ['0','1','2','3','4','5','6','7','8','9']
                    if line[j + 3] in tmp:
                        checksum += int(line[j + 3])
                    else:
                        checksum = 0
                        break
                if chr(checksum) == line[9]:
                    f = True
                    num = int(line[0])
                    minute = int(line[3])
                    second = int(line[4]) * 10 + int(line[5])
                    msecond = int(line[6]) * 100 + int(line[7]) * 10 + int(line[8])
                    #print(num,minute,second,msecond)
                    time[int(line[0])][turn[int(line[0])]] = round(minute * 60 + second + msecond / 1000, 3) #str(minute) + ':' + str(second) + '.' + str(msecond)
                    print(time[int(line[0])][turn[int(line[0])]])
                    confirmtime()
                    f = True
                else:
                    f = False
        if f:
            for i in range(30):
                ser.write('y'.encode())
    
    ser.close()
    root.after(1,inputserial)


port = search_com_port()

workbook = 'test.xlsx'
wb = openpyxl.load_workbook(workbook)
sheet = wb['Sheet1']
maxrow = [0, 0]
maxrow[0] = findmaxrow(0)
maxrow[1] = findmaxrow(1)
#print(maxrow)

time = []
for i in range(2):
    time.append([])
    for j in range(5):
        time[i].append('')
turn = [0, 0]

root = tkinter.Tk()
root.title("Talay1.0")
root.geometry("900x500")
canvas = tkinter.Canvas(root, width = 900, height = 500)
canvas.place(x=0,y=0)

name0label = tkinter.Label(root, text='Name')
name0label.grid(row=0, column=0, padx=5, pady=5)
name0box = tkinter.Entry(width=50)
name0box.grid(row=0, column=1, padx=5, pady=5)
name0entbutton = tkinter.Button(root, text='Save', command=lambda :nameinput(0))
name0entbutton.grid(row=0, column=2, padx=5, pady=5)

t10label = tkinter.Label(root, text='T1')
t10label.grid(row=1, column=0, padx=5, pady=5)
t10box = tkinter.Entry(width=50)
t10box.grid(row=1, column=1, padx=5, pady=5)
t10entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(0, 0))
t10entbutton.grid(row=1, column=2, padx=5, pady=5)

t20label = tkinter.Label(root, text='T2')
t20label.grid(row=2, column=0, padx=5, pady=5)
t20box = tkinter.Entry(width=50)
t20box.grid(row=2, column=1, padx=5, pady=5)
t20entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(1, 0))
t20entbutton.grid(row=2, column=2, padx=5, pady=5)

t30label = tkinter.Label(root, text='T3')
t30label.grid(row=3, column=0, padx=5, pady=5)
t30box = tkinter.Entry(width=50)
t30box.grid(row=3, column=1, padx=5, pady=5)
t30entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(2, 0))
t30entbutton.grid(row=3, column=2, padx=5, pady=5)

t40label = tkinter.Label(root, text='T4')
t40label.grid(row=4, column=0, padx=5, pady=5)
t40box = tkinter.Entry(width=50)
t40box.grid(row=4, column=1, padx=5, pady=5)
t40entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(3, 0))
t40entbutton.grid(row=4, column=2, padx=5, pady=5)

t50label = tkinter.Label(root, text='T5')
t50label.grid(row=5, column=0, padx=5, pady=5)
t50box = tkinter.Entry(width=50)
t50box.grid(row=5, column=1, padx=5, pady=5)
t50entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(4, 0))
t50entbutton.grid(row=5, column=2, padx=5, pady=5)

name1label = tkinter.Label(root, text='Name')
name1label.grid(row=0, column=3, padx=5, pady=5)
name1box = tkinter.Entry(width=50)
name1box.grid(row=0, column=4, padx=5, pady=5)
name1entbutton = tkinter.Button(root, text='Save', command=lambda :nameinput(1))
name1entbutton.grid(row=0, column=5, padx=5, pady=5)

t11label = tkinter.Label(root, text='T1')
t11label.grid(row=1, column=3, padx=5, pady=5)
t11box = tkinter.Entry(width=50)
t11box.grid(row=1, column=4, padx=5, pady=5)
t11entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(0, 1))
t11entbutton.grid(row=1, column=5, padx=5, pady=5)

t21label = tkinter.Label(root, text='T2')
t21label.grid(row=2, column=3, padx=5, pady=5)
t21box = tkinter.Entry(width=50)
t21box.grid(row=2, column=4, padx=5, pady=5)
t21entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(1, 1))
t21entbutton.grid(row=2, column=5, padx=5, pady=5)

t31label = tkinter.Label(root, text='T3')
t31label.grid(row=3, column=3, padx=5, pady=5)
t31box = tkinter.Entry(width=50)
t31box.grid(row=3, column=4, padx=5, pady=5)
t31entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(2, 1))
t31entbutton.grid(row=3, column=5, padx=5, pady=5)

t41label = tkinter.Label(root, text='T4')
t41label.grid(row=4, column=3, padx=5, pady=5)
t41box = tkinter.Entry(width=50)
t41box.grid(row=4, column=4, padx=5, pady=5)
t41entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(3, 1))
t41entbutton.grid(row=4, column=5, padx=5, pady=5)

t51label = tkinter.Label(root, text='T5')
t51label.grid(row=5, column=3, padx=5, pady=5)
t51box = tkinter.Entry(width=50)
t51box.grid(row=5, column=4, padx=5, pady=5)
t51entbutton = tkinter.Button(root, text='Save', command=lambda :timeinput(4, 1))
t51entbutton.grid(row=5, column=5, padx=5, pady=5)

resetbutton0 = tkinter.Button(root, text='Next Person', command=lambda :nextperson(0))
resetbutton0.grid(row=6, column=0, columnspan=3, sticky=tkinter.W+tkinter.E)

resetbutton1 = tkinter.Button(root, text='Next Person', command=lambda :nextperson(0))
resetbutton1.grid(row=6, column=3, columnspan=3, sticky=tkinter.W+tkinter.E)

root.after(1,inputserial)
root.mainloop()